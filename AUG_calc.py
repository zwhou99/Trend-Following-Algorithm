import numpy as np
import pandas as pd
import time
import argparse
import fcntl
import shutil
import openpyxl
import random
from datetime import datetime as dt
from numba import njit
import matplotlib.pyplot as plt
from math import ceil

# Function to format currency values
def format_currency(val):
    return "${:,.2f}".format(val)

# Function to format percentage values with 2 decimal places
def format_percentage(val):
    return "{:.2f}%".format(val * 100)

# Function to format percentage values with 4 decimal places
def format_percentage_4_decimal(val):
    return "{:.4f}%".format(val * 100)

# Function to format negative currency values
def format_negative_currency(val):
    return "(${:,.2f})".format(abs(val))

# Function to format Sharpe ratio values
def format_sharpe_ratio(val):
    return "{:.5f}".format(val)

@njit
def compute_positions(HH, LL, bHH, sLL, sellShorts, buyLongs, High, Low, Open, Close, L, S, start, end, bars_back,
                      PV, slpg, E, position):

    N = len(HH)  # Length of HH
    num_trades = 0  # Initialize the number of trades
    longtrades = 0  # Initialize the number of long trades
    shorttrades = 0  # Initialize the number of short trades
    trades = np.zeros(N)  # Initialize the trades array
    pnl = []  # Initialize the profit and loss list
    pnl_rate = []  # Initialize the profit and loss rate list
    trade_bars = []  # Initialize the trade bars list

    # Loop through the range of bars, starting from the maximum of bars_back+1 and start, to the end
    for k in range(max(bars_back + 1, start), end):

        traded = False  # Initialize the traded flag
        delta = PV * (Close[k] - Close[k - 1]) * position  # Calculate the change in value due to price change

        # Check if there is no position
        if position == 0:

            buy = buyLongs[k]  # Check for buy signal
            sell = sellShorts[k]  # Check for sell signal

            # If both buy and sell signals are present
            if buy and sell:
                # Calculate the change in value due to both buy and sell signals
                delta = -slpg + PV * (LL[k] - HH[k])
                pnl.append(delta)
                pnl_rate.append(delta / E[k - 1])
                trade_bars.append(1)
                longtrades += 0.5
                shorttrades += 0.5
                num_trades += 1
                trades[k] = 1
            else:
                # If there is only a buy signal
                if buy:
                    lastequity = E[k - 1]
                    lastprice = bHH[k]
                    lasttime = k
                    longtrades += 1
                    delta = -slpg / 2 + PV * (Close[k] - lastprice)
                    position = 1
                    traded = True
                    benchmarkLong = High[k]
                    trades[k] = 0.5
                    num_trades += 0.5
                # If there is only a sell signal
                elif sell:
                    shorttrades += 1
                    lastequity = E[k - 1]
                    lastprice = sLL[k]
                    lasttime = k
                    delta = -slpg / 2 - PV * (Close[k] - lastprice)
                    position = -1
                    traded = True
                    benchmarkShort = Low[k]
                    trades[k] = 0.5
                    num_trades += 0.5

        # If the position is long and not traded
        if position == 1 and not traded:
            sellShort = sellShorts[k]  # Check for sellShort signal
            sell = Low[k] <= (benchmarkLong * (1 - S))  # Check for sell signal

            if sellShort:
                pnl.append(-slpg + PV * (sLL[k] - lastprice))
                pnl_rate.append((-slpg + PV * (sLL[k] - lastprice)) / lastequity)
                trade_bars.append(k - lasttime + 1)
                lastequity = E[k - 1] - slpg / 2 + PV * (sLL[k] - Close[k - 1])
                lastprice = sLL[k]
                lasttime = k
                shorttrades += 1
                delta = delta - slpg - 2 * PV * (Close[k] - lastprice)
                position = -1
                benchmarkShort = Low[k]
                trades[k] = 1
                num_trades += 1
            elif sell:
                pnl.append(-slpg + PV * (min(Open[k], (benchmarkLong * (1 - S))) - lastprice))
                pnl_rate.append(
                    (-slpg + PV * (min(Open[k], (benchmarkLong * (1 - S))) - lastprice)) / lastequity)
                trade_bars.append(k - lasttime + 1)
                delta = delta - slpg / 2 - PV * (
                        Close[k] - min(Open[k], (benchmarkLong * (1 - S))))
                position = 0
                trades[k] = 0.5
                num_trades += 0.5

            benchmarkLong = max(High[k], benchmarkLong)

        # If the position is short and not traded
        if position == -1 and not traded:
            buyLong = buyLongs[k]  # Check for buyLong signal
            buy = High[k] >= (benchmarkShort * (1 + S))  # Check for buy signal

            if buyLong:
                longtrades += 1
                pnl.append(-slpg - PV * (bHH[k] - lastprice))
                pnl_rate.append((-slpg - PV * (bHH[k] - lastprice)) / lastequity)
                trade_bars.append(k - lasttime + 1)
                lastequity = E[k - 1] - slpg / 2 - PV * (bHH[k] - lastprice)
                lastprice = bHH[k]
                lasttime = k
                delta = delta - slpg + 2 * PV * (Close[k] - lastprice)
                position = 1
                benchmarkLong = High[k]
                trades[k] = 1
                num_trades += 1
            elif buy:
                pnl.append(-slpg - PV * (max(Open[k], (benchmarkShort * (1 + S))) - lastprice))
                pnl_rate.append((-slpg - PV * (max(Open[k], (benchmarkShort * (1 + S))) - lastprice)) / lastequity)
                trade_bars.append(k - lasttime + 1)
                delta = delta - slpg / 2 + PV * (Close[k] - max(Open[k], (benchmarkShort * (1 + S))))
                position = 0
                trades[k] = 0.5
                num_trades += 0.5

            benchmarkShort = min(Low[k], benchmarkShort)

        # Update equity
        E[k] = E[k - 1] + delta

    # If the position is not closed at the end, append the profit/loss information
    if position != 0:
        pnl.append(-slpg / 2 + PV * (Close[end-1] - lastprice) * position)
        pnl_rate.append((-slpg / 2 + PV * (Close[end-1] - lastprice) * position) / lastequity)
        trade_bars.append(end - lasttime)


    return E, trades, position, longtrades, shorttrades, pnl, pnl_rate, trade_bars


# Loop through years from 1 to 2
for y in range(1, 3):
    # Generate the filename for the Excel file for y years and 3 months
    filename = "AUG_74000/AUG_find_%dy_%dm.xlsx" % (y, 1)
    print(filename)

    # Load the Excel workbook and select the first sheet
    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    # Read data from the sheet starting from the third row and store it in a list
    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        data.append(row)

    # Sort the data based on the first column in descending order
    data.sort(key=lambda x: x[0], reverse=True)

    # Loop through months from 1 to 3
    for m in range(2, 4):
        # Generate the filename for the Excel file for y years and m months
        filename2 = "AUG_74000/AUG_find_%dy_%dm.xlsx" % (y, m)

        # Load the Excel workbook and select the first sheet
        wb = openpyxl.load_workbook(filename2)
        ws = wb[wb.sheetnames[0]]

        # Initialize a new list for storing the modified data
        data2 = []
        j = m
        # Loop through the data, and select specific rows based on the variable j
        for i in range(j - 1, len(data), j):
            data2.append((data[i][0], data[i][1], data[i][2], data[i - j + 1][3], data[i][4], data[i][5], data[i][6],
                          data[i][7], data[i][8]))

        # Sort the data2 based on the first column in ascending order
        data2.sort(key=lambda x: x[0])

        # Delete all rows from the third row onwards
        for row_idx in range(ws.max_row, 2, -1):
            ws.delete_rows(row_idx)

        # Write data2 to the sheet, starting from the third row
        for row_data in data2:
            ws.append(row_data)

        # Save the workbook
        wb.save(filename2)


# Set print options for NumPy arrays to suppress scientific notation
np.set_printoptions(suppress=True)

# Load data from a CSV file and preprocess it
data_file = 'AUG-5min.csv'
df = pd.read_csv(data_file)
df.Date = pd.to_datetime(df.Date)
df['Datetime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str))

# Initialize an array to store the profit-to-drawdown ratios for different year-month combinations
ratios = np.zeros((2, 3))

# Loop through years from 1 to 2 and months from 1 to 3
for year in range(1, 3):
    for month in range(1, 4):
        # Read and sort data from an Excel file
        filename = "AUG_74000/AUG_find_%dy_%dm.xlsx" % (year, month)
        wb = openpyxl.load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            data.append(row)
        data.sort(key=lambda x: x[0])

        # Initialize variables for trading
        E0 = 500000
        bars_back = 10001
        slpg = 70
        PV = 1000
        N = len(df)
        df["E"] = np.zeros(N) + E0
        num_trades = 0
        long_trades = 0
        short_trades = 0
        pnls = []
        pnl_rates = []
        trade_bars = []

        position = 0
        # Compute trading positions for each data point
        for id in range(0, len(data)):
            # Perform various calculations and update the DataFrame
            start = df.Date.searchsorted(data[id][2], side='left')
            end = df.Date.searchsorted(data[id][3], side='right')

            L = data[id][4]
            S = data[id][5]

            df["HH"] = df.High.rolling(L, closed='left').max()
            df["LL"] = df.Low.rolling(L, closed='left').min()
            df['bHH'] = df[['HH', 'Open']].max(axis=1)
            df['sLL'] = df[['LL', 'Open']].min(axis=1)
            df["sellShort"] = df.Low <= df.LL
            df["buyLong"] = df.High >= df.HH

            E, trades, position, longtrade, shorttrade, pnl, pnl_rate,\
            trade_bar = compute_positions(df.HH.values, df.LL.values, df.bHH.values, df.sLL.values, df.sellShort.values,
                                          df.buyLong.values, df.High.values, df.Low.values, df.Open.values,
                                          df.Close.values, L, S, start, end, bars_back, PV, slpg, df.E.values, 0)


            df.E = E
            num_trades += round(longtrade + shorttrade)
            long_trades += longtrade
            short_trades += shorttrade
            pnls.extend(pnl)
            pnl_rates.extend(pnl_rate)
            trade_bars.extend(trade_bar)

        # Calculate various performance metrics
        df["Emax"] = df.E.cummax()
        df["DD"] = df.E - df.Emax

        start = df.Date.searchsorted(data[0][2], side='left')
        profit = df.E[end-1] - df.E[0]
        maxdrawdown = abs(min(df.DD[0: end]))
        avgdrawdown = df.DD[start: end].mean()
        ratios[(year-1), (month-1)] = profit / maxdrawdown

        df['E_diff'] = df['E'].diff()
        bar_pnl = np.array(df.E_diff[start:end])

        pnls = np.array(pnls)
        trade_bars = np.array(trade_bars)
        positive_pnls = pnls[pnls > 0]
        negative_pnls = pnls[pnls < 0]

        positive_pnls_sum = positive_pnls.sum()
        negative_pnls_sum = negative_pnls.sum()
        positive_pnls_count = len(positive_pnls)
        negative_pnls_count = len(negative_pnls)

        positive_bars = trade_bars[pnls > 0]
        negative_bars = trade_bars[pnls < 0]

        pnl_rates = np.array(pnl_rates)
        positive_pnl_rates = pnl_rates[pnl_rates > 0]
        negative_pnl_rates = pnl_rates[pnl_rates < 0]
        positive_pnl_rates_avg = positive_pnl_rates.mean()
        negative_pnl_rates_avg = negative_pnl_rates.mean()
        avg_ror = pnl_rates.mean()
        avg_std = pnl_rates.std()

        period_pnl = np.array(df.E.diff(72)[end:start:-72])
        positive_period_pnl = period_pnl[period_pnl > 0]
        negative_period_pnl = period_pnl[period_pnl < 0]
        positive_period_pnl_sum = positive_period_pnl.sum()
        negative_period_pnl_sum = negative_period_pnl.sum()

        # Print performance metrics
        print("Time Series: AUG/AUG-5min.csv")
        print("In Sample: %d years" % year)
        print("Out of Sample: %d months" % month)
        print("----------------------")
        print("Net Equity", format_currency(df.E[end - 1]))
        print("Net Profit", format_currency(profit))
        print("Net Profit To Worst Drawdown", format_percentage(profit / maxdrawdown))
        print("Average Loser", format_percentage(negative_pnl_rates_avg))
        print("Average Winner", format_percentage(positive_pnl_rates_avg))
        print("Average Winner To Average Loser",
              format_percentage(positive_pnl_rates_avg / abs(negative_pnl_rates_avg)))
        print("Gross Gain", format_currency(positive_pnls_sum))
        print("Gross Loss", format_negative_currency(negative_pnls_sum))
        print("Gross Gain Period", format_currency(positive_period_pnl_sum))
        print("Gross Loss Period", format_negative_currency(negative_period_pnl_sum))
        print("Profit Factor", round(positive_pnls_sum / abs(negative_pnls_sum), 2))
        print("Profit Factor Period", round(positive_period_pnl_sum / abs(negative_period_pnl_sum), 2))
        print("Trade Count", num_trades)
        print("Long Trade Count", round(long_trades * 2) / 2)
        print("Short Trade Count", round(short_trades * 2) / 2)
        print("Average Drawdown", format_negative_currency(avgdrawdown))
        print("Worst Drawdown", format_negative_currency(maxdrawdown))
        print("Percent Winners", format_percentage(positive_pnls_count / num_trades))
        print("Percent Losers", format_percentage(negative_pnls_count / num_trades))
        print("Best Winner", format_percentage(max(positive_pnl_rates)))
        print("Worst Loser", format_percentage(min(negative_pnl_rates)))
        print("Best Winner To Worst Loser", format_percentage(max(positive_pnl_rates) / abs(min(negative_pnl_rates))))
        print("Winners Losers Ratio", format_percentage(positive_pnls_count / negative_pnls_count))
        print("Sharpe Ratio", format_sharpe_ratio(avg_ror / avg_std))
        print("Average RoR", format_percentage_4_decimal(avg_ror))
        print("Average Std Dev", format_percentage(avg_std))
        print("Avg Bars In Trade", round(trade_bars.mean(), 2))
        print("Avg Bars In Winning Trade", round(positive_bars.mean(), 2))
        print("Avg Bars In Losing Trade", round(negative_bars.mean(), 2))
        # print("Big Winner Count", )
        print("----------------------")
        print("Slippage", slpg)
        print("Max Bars Back", bars_back)
        print("Series Date Start", df.Datetime[start])
        print("Series Date End", df.Datetime[end - 1])
        print()
        print()
        print()
        print()
        print()
        print()
        
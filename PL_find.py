import numpy as np
import pandas as pd
import time
import argparse
import fcntl
import shutil
import openpyxl
import random
from datetime import datetime as dt
from getratio import getratio



# Set print options for NumPy arrays to suppress scientific notation
np.set_printoptions(suppress=True)

# Initialize argument parser and add arguments for id
parser = argparse.ArgumentParser()
parser.add_argument("--id", type=int, default=17)

# Parse the command line arguments
args = parser.parse_args()
id = args.id

# Calculate year and month based on id value
month = 3
if id < 58:
    year = 1
elif id < 112:
    id -= 58
    year = 2
elif id < 162:
    id -= 112
    year = 3
elif id < 208:
    id -= 162
    year = 4
elif id < 250:
    id -= 208
    year = 5
elif id < 288:
    id -= 250
    year = 6
elif id < 322:
    id -= 288
    year = 7
elif id < 352:
    id -= 322
    year = 8
elif id < 378:
    id -= 352
    year = 9

# Define Ls and Ss lists for calculations
Ls = list(range(100, 1000, 200))
Ss = list(range(20, 100, 20))

# Initialize an array to store ratios
ratios = np.empty((1001, 101))
ratios[:, :] = np.nan

# Initialize variables for finding the maximum ratio
allmax = -np.inf
argmax = [-1, -1]




# Variables initialization
data_file = 'PL-5min.csv'
end = pd.to_datetime("04/06/2023") - pd.tseries.offsets.DateOffset(years=year, months=month)
times_start = pd.date_range(end, "10/03/2007", freq=str(-month) + 'M').sort_values() + \
              pd.tseries.offsets.DateOffset(days=7) - pd.tseries.offsets.DateOffset(months=1)
times_end = times_start + pd.tseries.offsets.DateOffset(years=year, days=-1)

# Set constants for calculations
bars_back = 10001
slpg = 148
PV = 50
E0 = 100000

# Read data from CSV file and convert date column to datetime
df = pd.read_csv(data_file)
df.Date = pd.to_datetime(df.Date)

# Calculate indices for in-sample data
start = df.Date.searchsorted(times_start[id], side='left')
end = df.Date.searchsorted(times_end[id], side='right')


# Iterate through different L and S values to find the best ratio
for i in range(len(Ls) * len(Ss)):
    L = Ls[i // len(Ss)]
    S = Ss[i % len(Ss)]
    if np.isnan(ratios[L, S]):

        # Calculate the performance ratio for the current L and S values
        curmax, profit, maxdrawdown = getratio(df, L, S, start, end, bars_back, PV, slpg, E0)
        ratios[L, S] = curmax
        print("L=%d, S=%.3f: %.6f" % (L * 10, S / 1000, curmax))

        # Adjust L and S values based on their comparisons with neighboring cells in the ratios matrix
        cont = True
        while cont:
            cont = False
            if L > 50:
                if np.isnan(ratios[L - 1, S]):
                    ratio, profit, maxdrawdown = getratio(df, L - 1, S, start, end, bars_back, PV, slpg, E0)
                    ratios[L - 1, S] = ratio
                    print("L=%d, S=%.3f: %.6f" % (L * 10 - 10, S / 1000, ratio))
                    if ratio > curmax:
                        curmax = ratio
                        cont = True
                        dL = -1
                        dS = 0

            if S < 100:
                if np.isnan(ratios[L, S + 1]):
                    ratio, profit, maxdrawdown = getratio(df, L, S + 1, start, end, bars_back, PV, slpg, E0)
                    ratios[L, S + 1] = ratio
                    print("L=%d, S=%.3f: %.6f" % (L * 10, S / 1000 + 0.001, ratio))
                    if ratio > curmax:
                        curmax = ratio
                        cont = True
                        dL = 0
                        dS = 1

            if L < 1000:
                if np.isnan(ratios[L + 1, S]):
                    ratio, profit, maxdrawdown = getratio(df, L + 1, S, start, end, bars_back, PV, slpg, E0)
                    ratios[L + 1, S] = ratio
                    print("L=%d, S=%.3f: %.6f" % (L * 10 + 10, S / 1000, ratio))
                    if ratio > curmax:
                        curmax = ratio
                        cont = True
                        dL = 1
                        dS = 0

            if S > 4:
                if np.isnan(ratios[L, S - 1]):
                    ratio, profit, maxdrawdown = getratio(df, L, S - 1, start, end, bars_back, PV, slpg, E0)
                    ratios[L, S - 1] = ratio
                    print("L=%d, S=%.3f: %.6f" % (L * 10, S / 1000 - 0.001, ratio))
                    if ratio > curmax:
                        curmax = ratio
                        cont = True
                        dL = 0
                        dS = -1

            if cont:
                L = L + dL
                S = S + dS

        # Update the best performance ratio if needed
        if curmax > allmax:
            allmax = curmax
            argmax = [L, S]


# Fine-tune the search for the best ratio by exploring the neighborhood of the best ratio found so far
for L in range(max(50, argmax[0]-25), min(1000, argmax[0]+26)):
    for S in range(max(4, argmax[1]-5), min(100, argmax[1]+6)):
        if np.isnan(ratios[L, S]):
            ratio, profit, maxdrawdown = getratio(df, L, S, start, end, bars_back, PV, slpg, E0)
            ratios[L, S] = ratio
            print("L=%d, S=%.3f: %.6f" % (L * 10, S / 1000, ratio))
            if ratio > allmax:
                argmax = [L, S]
                allmax = ratio


# Print the best L and S values, ratio, profit, and max drawdown
print(argmax)
ratio, profit, maxdrawdown = getratio(df, argmax[0], argmax[1], start, end, bars_back, PV, slpg, E0)
print("ratio: %.4f" % ratio)
print("profit: %.2f" % profit)
print("maxdrawdown: %.2f" % maxdrawdown)


# Define the output filename
filename = "PL_find_%dy_%dm.xlsx" % (year, month)


# Lock the file for exclusive access
f = open(filename, "r")
fcntl.flock(f.fileno(), fcntl.LOCK_EX)
print("File locked")

# Create a backup copy of the file
now = dt.now()
localtime = now.strftime("%Y-%m-%d-%H-%M-%S-%f")
shutil.copyfile(filename, filename + "_" + localtime + ".xlsx")
print("File copied1")

# Load the workbook and select the first worksheet
wb = openpyxl.load_workbook(filename)
ws = wb[wb.sheetnames[0]]

# Prepare new data for appending and sorting
new_data = (str(times_start[id]), str(times_end[id]), str(times_end[id] + pd.tseries.offsets.DateOffset(days=1)), str(times_end[id] + pd.tseries.offsets.DateOffset(months=month)), argmax[0] * 10, argmax[1] / 1000, ratio, profit, maxdrawdown)
data = [new_data]

# Load existing data, excluding the first two rows
for row in ws.iter_rows(min_row=3, values_only=True):
    data.append(row)

# Sort the data by date
data.sort(key=lambda x: x[0])

# Clear the original data, excluding the first two rows
for row in ws['A3:J' + str(ws.max_row)]:
    for cell in row:
        cell.value = None

# Write the sorted data back to the worksheet, starting from the third row
for i, row_data in enumerate(data, start=3):
    for j, cell_data in enumerate(row_data):
        ws.cell(row=i, column=j + 1, value=cell_data)

# Create a second backup copy of the file
shutil.copyfile(filename, filename + "_" + localtime + ".xlsx")
print("File copied2")

# Save the file, unlock it, and close it
wb.save(filename)
fcntl.flock(f.fileno(), fcntl.LOCK_UN)
f.close()
print("Successfully saved")